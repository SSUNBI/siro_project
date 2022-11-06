from this import d
import openpyxl
import calendar

def week_calculate(wb, name, week_end_day):
    cs = wb["Sheet1"]
    week_list = list()
    week_sum = 0
    now_date = cs["A2"].value
    now_year = now_date // 10000 + 2000
    now_month = (now_date % 10000) // 100
    now_day = now_date % 100
    weekday = calendar.weekday(now_year, now_month, now_day)
    month_range = calendar.monthrange(now_year, now_month)[1]
    cnt = 2
    temp_date = now_date
    temp_year = temp_date // 10000 + 2000
    temp_month = (temp_date % 10000) // 100
    temp_day = now_date % 100
    start_date = temp_date
    while True:
        if cnt >= cs.max_row:
            if weekday != week_end_day: #week_end_day ddnayo : 6
                week_list.append([start_date, temp_date, week_sum])
            break
        now_date = cs["A" + str(cnt)].value
        if temp_date == now_date:
            now_year = now_date // 10000 + 2000
            now_month = (now_date % 10000) // 100
            now_day = now_date % 100
            month_range = calendar.monthrange(now_year, now_month)[1]
            if cs["D" + str(cnt)].value == name:
                week_sum += cs["C" + str(cnt)].value
            cnt += 1
        else:
            while temp_date < now_date:
                end_date = temp_date
                temp_day += 1
                if temp_day > month_range:
                    temp_day = 1
                    temp_month += 1
                    if temp_month > 12:
                        temp_month = 1
                        temp_year += 1
                temp_date = (temp_year - 2000) * 10000 + temp_month * 100 + temp_day
                weekday = calendar.weekday(temp_year, temp_month, temp_day)
                if weekday == (week_end_day + 1) % 7:
                    break
            if weekday == (week_end_day + 1) % 7: #ddnayo
                week_list.append([start_date, end_date, week_sum])
                start_date = temp_date
                week_sum = 0
    print(week_list)
    return week_list

def add_calculate(wb, year, month):
    #야놀자
    ws = wb["yanolza"]

    #여기어때 시트 정산금액 텍스트 형식 숫자형식으로 W열에 저장
    sheet_yogi_A = wb["yogi_A"]
    for i in range(4,105):
        sheet_yogi_A.cell(row = i, column = 23, value = '=I'+str(i)+'*1')

    sheet_yogi_B = wb["yogi_B"]
    for i in range(4,105):
        sheet_yogi_B.cell(row = i, column = 23, value = '=I'+str(i)+'*1')

    sheet_yogi_C = wb["yogi_C"]
    for i in range(4,105):
        sheet_yogi_C.cell(row = i, column = 23, value = '=I'+str(i)+'*1')

    #주간 정산금액 불러오기
    #A_bank (야놀자 모든정산금액 + 떠나요 M+A동 + 여기어때 M+A동) - C,E,H열 데이터 뽑아오기
    #B_bank (떠나요B동 + 여기어때B동)
    #C_bank (떠나요C동 + 여기어때C동)
    #야놀자는 '야놀자펜션0117-'라는 이름으로 들어옴
    #떠나요는 '주식회사떠나요'라는 이름으로 들어옴
    #여기어때는 '호텔타임'라는 이름으로 들어옴
    #떠나요 계산기 일~토 화요일정산
    #여기어때 월~일 수요일정산
    #야놀자 월~일 목요일정산
    #mon : 0 tue = 1 wed = 2 ...

    ddn_list = week_calculate(wb, "ddnayo", 5)
    yanolza_list = week_calculate(wb, "yanolza", 6)
    yogi_A_list = week_calculate(wb, "yogi_A", 6)
    yogi_B_list = week_calculate(wb, "yogi_B", 6)
    yogi_C_list = week_calculate(wb, "yogi_C", 6)

    sheet_A_bank = wb["A_bank"]
    sheet_B_bank = wb["B_bank"]
    sheet_C_bank = wb["C_bank"]
    yanolza_bank_list = []
    ddn_bank_M_A_list = []
    yogi_bank_M_A_list = []
    ddn_bank_B_list = []
    ddn_bank_C_list = []
    yogi_bank_B_list = []
    yogi_bank_C_list = []
    ddn_sum_list = []
    cnt = 0
    for row in sheet_A_bank.rows:
        if cnt < 8:
            cnt += 1
            continue
        sender = str(row[7].value)
        if sender == '주식회사떠나요':
            year_bank = int(row[2].value[:4])
            month_bank = int(row[2].value[5:7])
            date_bank = int(row[2].value[8:10])
            money_bank = int(row[4].value)
            settlement_date = (year_bank-2000)*10000 + (month_bank)*100 + (date_bank)
            ddn_bank_M_A_list.append([settlement_date, money_bank, sender])
        elif sender == '호텔타임':
            year_bank = int(row[2].value[:4])
            month_bank = int(row[2].value[5:7])
            date_bank = int(row[2].value[8:10])
            money_bank = int(row[4].value)
            settlement_date = (year_bank-2000)*10000 + (month_bank)*100 + (date_bank)
            yogi_bank_M_A_list.append([settlement_date, money_bank, sender])
        elif sender[:5] == '야놀자펜션':
            year_bank = int(row[2].value[:4])
            month_bank = int(row[2].value[5:7])
            date_bank = int(row[2].value[8:10])
            money_bank = int(row[4].value)
            settlement_date = (year_bank-2000)*10000 + (month_bank)*100 + (date_bank)
            yanolza_bank_list.append([settlement_date, money_bank, sender])
    cnt = 0
    for row in sheet_B_bank.rows:
        if cnt < 8:
            cnt += 1
            continue
        sender = str(row[7].value)
        if sender == '주식회사떠나요':
            year_bank = int(row[2].value[:4])
            month_bank = int(row[2].value[5:7])
            date_bank = int(row[2].value[8:10])
            money_bank = int(row[4].value)
            settlement_date = (year_bank-2000)*10000 + (month_bank)*100 + (date_bank)
            ddn_bank_B_list.append([settlement_date, money_bank, sender])
        elif sender == '호텔타임':
            year_bank = int(row[2].value[:4])
            month_bank = int(row[2].value[5:7])
            date_bank = int(row[2].value[8:10])
            money_bank = int(row[4].value)
            settlement_date = (year_bank-2000)*10000 + (month_bank)*100 + (date_bank)
            yogi_bank_B_list.append([settlement_date, money_bank, sender])
    cnt = 0
    for row in sheet_C_bank.rows:
        if cnt < 8:
            cnt += 1
            continue
        sender = str(row[7].value)
        if sender == '주식회사떠나요':
            year_bank = int(row[2].value[:4])
            month_bank = int(row[2].value[5:7])
            date_bank = int(row[2].value[8:10])
            money_bank = int(row[4].value)
            settlement_date = (year_bank-2000)*10000 + (month_bank)*100 + (date_bank)
            ddn_bank_C_list.append([settlement_date, money_bank, sender])
        elif sender == '호텔타임':
            year_bank = int(row[2].value[:4])
            month_bank = int(row[2].value[5:7])
            date_bank = int(row[2].value[8:10])
            money_bank = int(row[4].value)
            settlement_date = (year_bank-2000)*10000 + (month_bank)*100 + (date_bank)
            yogi_bank_C_list.append([settlement_date, money_bank, sender])
    

    #print(ddn_bank_M_A_list)
    #print(yogi_bank_M_A_list)
    #print(yanolza_bank_list)
    #print(ddn_bank_B_list)
    #print(ddn_bank_C_list)
    #print(yogi_bank_B_list)
    #print(yogi_bank_C_list)
    ddn_bank_M_A_list.reverse()#내림차순 정렬
    yanolza_bank_list.reverse()
    yogi_bank_M_A_list.reverse()
    ddn_bank_B_list.reverse()
    ddn_bank_C_list.reverse()
    yogi_bank_B_list.reverse()
    yogi_bank_C_list.reverse()
    
    for i in range(len(ddn_bank_M_A_list)):
        ddn_sum_list.append(ddn_bank_M_A_list[i][1] + ddn_bank_B_list[i][1] + ddn_bank_C_list[i][1])
    print(ddn_sum_list)

    #수식 적는 곳 시작
    sheet = wb["calculate"]

    sheet['A1'] = '매출종합본'
    sheet['B1'] = year
    sheet['C1'] = month

    sheet['A3'] = '회사명'
    sheet['A4'] = '떠나요'
    sheet['A5'] = '여기어때A동,M동'
    sheet['A6'] = '여기어때B동'
    sheet['A7'] = '여기어때C동'
    sheet['A8'] = '야놀자'
    sheet['A9'] = '총합'

    sheet['B3'] = '월매출총합'
    sheet['B4'] = '=SUM(ddnayo!J:J)'
    sheet['B5'] = '=SUM(yogi_A!W4:W105)'
    sheet['B6'] = '=SUM(yogi_B!W4:W105)'
    sheet['B7'] = '=SUM(yogi_C!W4:W105)'
    sheet['B8'] = '=SUM(yanolza!L:L)'
    sheet['B9'] = '=SUM(B4:B8)'   

    #떠나요 사이트 및 은행 정산현황
    sheet['A11'] = '떠나요'
    sheet['A12'] = '사이트'
    sheet['A13'] = '기간'
    sheet['B13'] = '금액'
    sheet['C12'] = '은행'
    sheet['C13'] = '정산일'
    sheet['D13'] = 'M+A'
    sheet['E13'] = 'B'
    sheet['F13'] = 'C'
    start = 14
    for i in ddn_list:
        sheet['A' + str(start)] = str(i[0]) + ' ~ ' + str(i[1])
        sheet['B' + str(start)] = i[2]
        for j in ddn_bank_M_A_list:
            if j[0] > i[1] and j[0] < i[1]+9:
                sheet['C'+str(start)] = j[0]
                sheet['D'+str(start)] = j[1]
        for q in ddn_bank_B_list:
            if q[0] > i[1] and q[0] < i[1]+9:
                sheet['e'+str(start)] = q[1]
        for k in ddn_bank_C_list:
            if k[0] > i[1] and k[0] < i[1]+9:
                sheet['f'+str(start)] = k[1]
        start += 1

    #야놀자 사이트 및 은행 정산현황
    sheet['A22'] = '야놀자'
    sheet['A23'] = '사이트'
    sheet['A24'] = '기간'
    sheet['B24'] = '금액'
    sheet['C23'] = '은행'
    sheet['C24'] = '정산일'
    sheet['D24'] = '금액'
    start = 25
    for i in yanolza_list:
        sheet['A' + str(start)] = str(i[0]) + ' ~ ' + str(i[1])
        sheet['B' + str(start)] = i[2]
        for j in yanolza_bank_list:
            if j[0] > i[1] and j[0] < i[1]+9:
                sheet['C'+str(start)] = j[0]
                sheet['D'+str(start)] = j[1]
        start += 1
    
    #여기어때 M+A 동 사이트 및 은행 정산 현황
    sheet['A33'] = '여기어때'
    sheet['A34'] = 'M+A동'
    sheet['A35'] = '사이트'
    sheet['A36'] = '기간'
    sheet['B24'] = '금액'
    sheet['C35'] = '은행'
    sheet['C36'] = '정산일'
    sheet['D36'] = '금액'
    start = 37
    for i in yogi_A_list:
        sheet['A' + str(start)] = str(i[0]) + ' ~ ' + str(i[1])
        sheet['B' + str(start)] = i[2]
        for j in yogi_bank_M_A_list:
            if j[0] > i[1] and j[0] < i[1]+9:
                sheet['C'+str(start)] = j[0]
                sheet['D'+str(start)] = j[1]
        start += 1

    #여기어때 B동 사이트 및 은행 정산 현황
    sheet['E34'] = 'B동'
    sheet['E35'] = '사이트'
    sheet['E36'] = '기간'
    sheet['F36'] = '금액'
    sheet['G35'] = '은행'
    sheet['G35'] = '정산일'
    sheet['H36'] = '금액'
    start = 37
    for i in yogi_B_list:
        sheet['E' + str(start)] = str(i[0]) + ' ~ ' + str(i[1])
        sheet['F' + str(start)] = i[2]
        for j in yogi_bank_B_list:
            if j[0] > i[1] and j[0] < i[1]+9:
                sheet['G'+str(start)] = j[0]
                sheet['H'+str(start)] = j[1]
        start += 1

    #여기어때 C동 사이트 및 은행 정산 현황
    sheet['I34'] = 'B동'
    sheet['I35'] = '사이트'
    sheet['I36'] = '기간'
    sheet['J36'] = '금액'
    sheet['K35'] = '은행'
    sheet['K35'] = '정산일'
    sheet['L36'] = '금액'
    start = 37
    for i in yogi_C_list:
        sheet['I' + str(start)] = str(i[0]) + ' ~ ' + str(i[1])
        sheet['J' + str(start)] = i[2]
        for j in yogi_bank_C_list:
            if j[0] > i[1] and j[0] < i[1]+9:
                sheet['K'+str(start)] = j[0]
                sheet['L'+str(start)] = j[1]
        start += 1

#시트 autofit
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
        adjusted_width = (max_length + 2)*1.2
        sheet.column_dimensions[column].width = adjusted_width